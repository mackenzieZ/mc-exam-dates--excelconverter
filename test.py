#Project Name: MCcalendarDates
#Project Description: This program's purpose is to take the course information and exam dates from an excel file and generating an .ics calendar file 
# that can be imported into any calendar service such as GoogleCalendar or Outlook
#Creator: Mackenzie Zappe
#Date started: 10/27/2020
#Date last updated: 10/29/2020

#openpyxl is an open source library that method in place to read from an .xsxl file into a python program
import openpyxl

wb_obj = openpyxl.load_workbook("testExcelSheet.xlsx") 
 
sheet_obj = wb_obj.active

#various list declarations to be used in program
#the c then number lists refer to different lists that are used to manage/manipulate the exams for a specific course
c2 = []
c3 = []
c4 = []
c5 = []
#the course lists refer to the describable information of the course (ie. professor, section, class name)
course = []
course2 = []

#determines how many courses are being offered this semester 
numCourses = sheet_obj.max_row

#used to read all the information from the excel sheet into respective lists
#### Note the ranges for the loops, this will differ based on the format of the excel sheet that you use ####
for l in range(numCourses):
    #specifically reads the exam information from the file to the c lists
    for j in range(4 , sheet_obj.max_column +1):
            c2 = sheet_obj.cell(row = 1, column = j)
            c3.append(c2.value.year)
            c3.append(c2.value.month)
            c3.append(c2.value.day)
            c4 = str(c3[0]) + str(c3[1]) + str(c3[2])
            c5.append(c4)
    #specifically reads the course information from the file into the course lists
    for k in range(1, 4):
        course = sheet_obj.cell(row = 1, column = k)
        course2.append(course.value)

#opens/creates the .ics file
f = open("ExamDateCalendarEntries.ics", "w")

#begins writing to the .ics file 
### Note about formatting .ics file - cannot have spaces after : and must have \n end the line
f.write("BEGIN:VCALENDAR\n")
f.write("VERSION:2.0\n")
#loops through the number of courses that are in gthe file
for x in range(numCourses):
    #loops through the number of exams each course has scheduledS
    for y in range(len(c5)):
        f.write("BEGIN:VEVENT\n")
        f.write("DTSTART:" + str(c5[y]) + "\n")
        c5[y] = str(int(c5[y]) + 1)
        f.write("DTEND:" + str(c5[y]) + "\n")
        f.write("SUMMARY:" + str(course2[0]) + " " + str(course2[1]) + " " + str(course2[2]) + " Exam " + str(y+1) + "\n")
        f.write("END:VEVENT\n")
f.write("END:VCALENDAR\n")

f.close()
